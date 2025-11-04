"""
Generator danych treningowych od modelu nauczyciela.
Generuje różnorodne dialogi NPC z wysoką temperaturą (~2.0).
Pobiera prompty bezpośrednio z pliku Excel (prompty.xlsx).
"""

import torch
import yaml
import json
import random
from transformers import AutoModelForCausalLM, AutoTokenizer
from tqdm import tqdm
from pathlib import Path
from typing import List, Dict
from openpyxl import load_workbook


class NPCDataGenerator:
    """Generuje dane treningowe używając dużego modelu jako nauczyciela."""
    
    def __init__(self, config_path: str = "config.yaml"):
        """Inicjalizacja generatora."""
        with open(config_path, 'r') as f:
            self.config = yaml.safe_load(f)
        
        self.teacher_config = self.config['teacher_model']
        self.data_config = self.config['data_generation']
        
        print(f"Ładowanie modelu nauczyciela: {self.teacher_config['name']}")
        self.load_teacher_model()
        
        # Załaduj prompty z pliku Excel
        self.prompts = self._load_prompts_from_excel()
        
    def _load_prompts_from_excel(self) -> List[str]:
        """Ładuje wszystkie prompty z pliku prompty.xlsx."""
        excel_path = Path(__file__).parent.parent / "prompty.xlsx"
        
        if not excel_path.exists():
            print(f"⚠️  Plik {excel_path} nie istnieje!")
            raise FileNotFoundError(f"Brak pliku {excel_path}")
        
        wb = load_workbook(excel_path)
        all_prompts = []
        
        print(f"\n📊 Ładowanie promptów z Excel...")
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheet_prompts = []
            
            # Zbierz wszystkie niepuste komórki z arkusza
            for row in ws.rows:
                for cell in row:
                    if cell.value and str(cell.value).strip():
                        content = str(cell.value).strip()
                        # Pomiń nagłówki typu "Input", "Memories"
                        if content not in ['Input', 'Memories', 'input', 'memories', 'INPUT', 'MEMORIES']:
                            sheet_prompts.append(content)
            
            if sheet_prompts:
                all_prompts.extend(sheet_prompts)
                print(f"  ✓ {sheet_name}: {len(sheet_prompts)} promptów")
        
        print(f"\n✅ Załadowano łącznie {len(all_prompts)} promptów z {len(wb.sheetnames)} arkuszy\n")
        return all_prompts
        
    def load_teacher_model(self):
        """Ładuje model nauczyciela z optymalizacjami."""
        self.tokenizer = AutoTokenizer.from_pretrained(
            self.teacher_config['name'],
            trust_remote_code=True
        )
        
        # Ustawienie pad_token jeśli nie istnieje
        if self.tokenizer.pad_token is None:
            self.tokenizer.pad_token = self.tokenizer.eos_token
        
        # Ładowanie modelu z opcjonalną kwantyzacją
        model_kwargs = {
            "device_map": "auto",
            "torch_dtype": torch.float16,
        }
        
        if self.teacher_config.get('load_in_8bit', False):
            model_kwargs['load_in_8bit'] = True
        
        self.model = AutoModelForCausalLM.from_pretrained(
            self.teacher_config['name'],
            **model_kwargs
        )
        self.model.eval()
    
    @torch.no_grad()
    def generate_response(self, prompt: str) -> str:
        """Generuje odpowiedź od modelu nauczyciela z wysoką temperaturą."""
        try:
            # Prompty z Excel mogą być długie - zwiększamy max_length
            inputs = self.tokenizer(
                prompt,
                return_tensors="pt",
                padding=True,
                truncation=True,
                max_length=2048  # Zwiększone dla dłuższych promptów z Excel
            ).to(self.model.device)
            
            # FIX: Bezpieczne parametry aby uniknąć CUDA assert errors
            safe_temperature = min(self.teacher_config.get('temperature', 1.0), 1.2)
            
            outputs = self.model.generate(
                input_ids=inputs['input_ids'],
                attention_mask=inputs['attention_mask'],
                max_new_tokens=self.teacher_config.get('max_new_tokens', 128),
                temperature=safe_temperature,
                top_p=min(self.teacher_config.get('top_p', 0.9), 0.9),
                top_k=50,
                do_sample=True,
                pad_token_id=self.tokenizer.pad_token_id,
                eos_token_id=self.tokenizer.eos_token_id,
                repetition_penalty=1.1,
                renormalize_logits=True,  # Zapobiega NaN
                no_repeat_ngram_size=3,  # Zapobiega zapętleniom
            )
        except RuntimeError as e:
            if "CUDA" in str(e) or "assert" in str(e):
                print(f"⚠️  CUDA error, fallback to greedy decoding...")
                # Fallback: bezpieczna generacja bez sampling
                outputs = self.model.generate(
                    input_ids=inputs['input_ids'],
                    attention_mask=inputs['attention_mask'],
                    max_new_tokens=min(self.teacher_config.get('max_new_tokens', 128), 64),
                    do_sample=False,  # Greedy
                    pad_token_id=self.tokenizer.pad_token_id,
                    eos_token_id=self.tokenizer.eos_token_id,
                )
            else:
                raise
        
        # Dekodowanie tylko wygenerowanej części
        generated_text = self.tokenizer.decode(
            outputs[0][inputs['input_ids'].shape[1]:],
            skip_special_tokens=True
        )
        
        return generated_text.strip()
    
    def generate_dataset(self, output_path: str = None):
        """Generuje kompletny dataset treningowy z promptów Excel."""
        # Użyj ścieżki z configu jeśli nie podano jawnie
        if output_path is None:
            output_dir = self.data_config.get('output_dir', 'data')
            output_file = self.data_config.get('output_file', 'teacher_dataset.jsonl')
            output_path = Path(output_dir) / output_file
        else:
            output_path = Path(output_path)
            
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        num_samples = min(self.data_config['num_samples'], len(self.prompts))
        
        if num_samples < self.data_config['num_samples']:
            print(f"⚠️  Mam tylko {len(self.prompts)} promptów, generuję {num_samples} przykładów")
        
        # Losowo wybierz prompty jeśli jest ich więcej niż potrzeba
        selected_prompts = random.sample(self.prompts, num_samples) if len(self.prompts) > num_samples else self.prompts
        
        print(f"Generowanie {num_samples} przykładów z promptów Excel...")
        
        with open(output_path, 'w', encoding='utf-8') as f:
            for i, prompt in enumerate(tqdm(selected_prompts)):
                # Każdy prompt z Excel jest kompletnym inputem
                response = self.generate_response(prompt)
                
                # Zapisz parę prompt-odpowiedź
                example = {
                    'prompt': prompt,
                    'response': response,
                    'id': i
                }
                
                f.write(json.dumps(example, ensure_ascii=False) + '\n')
                
                # Opcjonalne opóźnienie żeby nie przegrzać GPU
                if i % 50 == 0:
                    torch.cuda.empty_cache()
        
        print(f"✓ Dataset zapisany do {output_path}")
        print(f"✓ Wygenerowano {num_samples} przykładów")


def main():
    """Główna funkcja generująca dane."""
    import argparse
    
    parser = argparse.ArgumentParser(description='Generowanie danych od nauczyciela')
    parser.add_argument(
        '--config',
        type=str,
        default='config.yaml',
        help='Ścieżka do pliku konfiguracyjnego'
    )
    
    args = parser.parse_args()
    
    generator = NPCDataGenerator(config_path=args.config)
    generator.generate_dataset()


if __name__ == "__main__":
    main()
